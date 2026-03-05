"""
Bulk import of users via Excel.

Used by:
  GET  /admin/users/import-template
  POST /admin/users/import
  GET  /admin/users/export
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from bulk_import import HEADER_FILL, HEADER_FONT, EXAMPLE_FILL

USER_HEADERS = ["operacao", "email", "nome", "perfil", "ativo"]

USER_EXAMPLE_ROW_1 = [
    "C",
    "novo@empresa.com",
    "Novo Usuário",
    "SOLICITANTE",
    "Sim",
]

USER_EXAMPLE_ROW_2 = [
    "E",
    "existente@empresa.com",
    "Nome Atualizado",
    "TRIAGEM",
    "Sim",
]

VALID_OPS = ("C", "E")
VALID_PERFIS = ("ADMIN", "MASTER", "TRIAGEM", "FISCAL", "MRP", "SOLICITANTE")
VALID_ATIVO = ("Sim", "Não")


def build_user_template_xlsx() -> BytesIO:
    """Build Excel template with Usuários sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuários"

    for col_idx, header in enumerate(USER_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, val in enumerate(USER_EXAMPLE_ROW_1, start=1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.fill = EXAMPLE_FILL
    for col_idx, val in enumerate(USER_EXAMPLE_ROW_2, start=1):
        cell = ws.cell(row=3, column=col_idx, value=val)
        cell.fill = EXAMPLE_FILL

    for row in range(4, 12):
        for col_idx in range(1, len(USER_HEADERS) + 1):
            ws.cell(row=row, column=col_idx, value="")

    for col_idx in range(1, len(USER_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            12, len(str(USER_HEADERS[col_idx - 1])) + 2
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_user_export_xlsx(users: list[dict[str, Any]]) -> BytesIO:
    """Build Excel export with real user data. NEVER exports passwords."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuários"

    for col_idx, header in enumerate(USER_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, u in enumerate(users, start=2):
        ws.cell(row=row_idx, column=1, value="E")
        ws.cell(row=row_idx, column=2, value=u.get("email") or "")
        ws.cell(row=row_idx, column=3, value=u.get("name") or "")
        ws.cell(row=row_idx, column=4, value=u.get("role_name") or "")
        ws.cell(row=row_idx, column=5, value="Sim" if u.get("is_active", True) else "Não")

    for col_idx in range(1, len(USER_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, 18)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _cell_value(cell) -> str | int | float | None:
    val = cell.value
    if val is None:
        return None
    if isinstance(val, str) and val.strip() == "":
        return None
    return val


def _is_valid_email(s: str) -> bool:
    """Basic check: contains @ and at least one dot after it."""
    if not s or "@" not in s:
        return False
    local, rest = s.split("@", 1)
    return bool(local.strip() and "." in rest)


def parse_and_validate_user_excel(file_content: bytes, db) -> dict:
    """
    Parse and validate user Excel file. Returns result for Usuários sheet.
    Does NOT persist — caller handles dry_run and commit.
    """
    from orm_models import UserORM, RoleORM

    wb = load_workbook(BytesIO(file_content), read_only=False, data_only=True)
    emails_in_db = {r.email.lower() for r in db.query(UserORM).all()}
    role_by_name: dict[str, int] = {
        (r.name or "").upper(): r.id for r in db.query(RoleORM).all()
    }

    result: dict[str, Any] = {"dry_run": True, "users": None}

    if "Usuários" not in wb.sheetnames:
        result["_error"] = "Aba 'Usuários' não encontrada no arquivo."
        return result

    ws = wb["Usuários"]
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if not header_row:
        result["_error"] = "Cabeçalho da aba Usuários não encontrado."
        return result

    headers = [str(h).strip().lower() if h else "" for h in header_row[0]]
    col_map: dict[str, int] = {}
    for idx, h in enumerate(headers):
        if h:
            col_map[h] = idx + 1

    for required in ["operacao", "email", "nome"]:
        if required not in col_map:
            result["_error"] = f"Coluna obrigatória '{required}' não encontrada na aba Usuários."
            return result

    rows_result: list[dict] = []
    total = 0
    valid = 0
    error_count = 0
    warning_count = 0

    for row_num in range(2, ws.max_row + 1):
        row_data: dict[str, Any] = {}
        for col_name, col_idx in col_map.items():
            cell = ws.cell(row=row_num, column=col_idx)
            row_data[col_name] = _cell_value(cell)

        if all(v is None for v in row_data.values()):
            continue

        total += 1
        errors: list[str] = []
        warnings: list[str] = []
        op = str(row_data.get("operacao", "") or "").strip().upper()
        email_raw = row_data.get("email")
        email = str(email_raw or "").strip().lower() if email_raw else ""
        nome = str(row_data.get("nome") or "").strip()
        perfil_raw = str(row_data.get("perfil") or "").strip().upper()
        ativo_raw = str(row_data.get("ativo") or "").strip()

        if op not in VALID_OPS:
            errors.append("operacao deve ser 'C' (Criar) ou 'E' (Editar)")
            rows_result.append({
                "row_number": row_num,
                "operacao": op or "?",
                "email": email or None,
                "nome": nome or None,
                "perfil": perfil_raw or None,
                "ativo": ativo_raw or None,
                "status": "error",
                "errors": errors,
                "warnings": warnings,
                "data": row_data,
            })
            error_count += 1
            continue

        if not email:
            errors.append("email é obrigatório")

        if email and not _is_valid_email(email):
            warnings.append("email deve conter @ e domínio (ex: usuario@empresa.com)")

        if op == "E" and email and email not in emails_in_db:
            errors.append(
                f"Usuário com e-mail '{email}' não encontrado. Para criar use operacao=C."
            )

        if op == "C" and email and email in emails_in_db:
            errors.append(f"E-mail '{email}' já cadastrado. Use operacao=E para editar.")

        if perfil_raw:
            if perfil_raw not in VALID_PERFIS:
                errors.append(
                    f"perfil deve ser um de: {', '.join(VALID_PERFIS)}"
                )
        elif op == "C":
            errors.append("perfil é obrigatório para operacao C (Criar)")

        if ativo_raw and ativo_raw not in VALID_ATIVO:
            warnings.append("ativo deve ser 'Sim' ou 'Não'")

        status_out = "error" if errors else ("warning" if warnings else "ok")
        if errors:
            error_count += 1
        elif warnings:
            warning_count += 1
        else:
            valid += 1

        rows_result.append({
            "row_number": row_num,
            "operacao": op,
            "email": email or None,
            "nome": nome or None,
            "perfil": perfil_raw or None,
            "ativo": ativo_raw or None,
            "status": status_out,
            "errors": errors,
            "warnings": warnings,
            "data": row_data,
        })

    result["users"] = {
        "total_rows": total,
        "valid_rows": valid,
        "error_rows": error_count,
        "warning_rows": warning_count,
        "rows": rows_result,
    }

    return result
