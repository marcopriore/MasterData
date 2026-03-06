"""
Bulk import of materials via Excel.

Used by:
  GET  /api/database/materials/import-template
  POST /api/database/materials/import
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Colunas do template (ordem exata)
TEMPLATE_HEADERS = [
    "operacao",
    "codigo_material",
    "descricao",
    "pdm_code",
    "status",
    "unit_of_measure",
    "material_type",
    "industry_sector",
    "base_unit",
    "gross_weight",
    "net_weight",
    "weight_unit",
    "volume",
    "volume_unit",
    "lead_time",
    "min_stock",
    "max_stock",
    "standard_price",
    "price_unit",
    "currency",
    "ncm",
    "cfop",
    "origem",
]

# Mapeamento coluna template -> campo MaterialDatabaseORM (apenas campos que existem no modelo)
COL_TO_MODEL: dict[str, str] = {
    "descricao": "description",
    "pdm_code": "pdm_code",
    "status": "status",
    "unit_of_measure": "unit_of_measure",
    "material_type": "material_type",
    "gross_weight": "gross_weight",
    "net_weight": "net_weight",
    "lead_time": "lead_time",
    "min_stock": "min_stock",
    "max_stock": "max_stock",
    "standard_price": "standard_price",
    "ncm": "ncm",
    "cfop": "cfop",
    "origem": "origin",
}
# codigo_material mapeia para id_erp na edição

HEADER_FILL = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
EXAMPLE_FILL = PatternFill(start_color="fffde7", end_color="fffde7", fill_type="solid")

EXAMPLE_ROW = [
    "C",
    "",
    "ROLAMENTO ESFERAS 6205 25MM",
    "PDM-ROL-001",
    "Ativo",
    "UN",
    "Rolamento",
    "",
    "",
    0.25,
    0.23,
    "KG",
    "",
    "",
    14,
    10.0,
    100.0,
    45.90,
    "",
    "BRL",
    "8482.10.10",
    "6102",
    "0 - Nacional",
]

INSTRUCTIONS = [
    "operacao aceita apenas C (Criar) ou E (Editar)",
    "Para operacao=E, codigo_material é obrigatório",
    "Para operacao=C, codigo_material será gerado automaticamente",
    "Campos obrigatórios para C: descricao, pdm_code",
    "status aceita: Ativo, Bloqueado, Obsoleto",
    "Não alterar os nomes das colunas do cabeçalho",
]


def _material_to_export_row(m: dict[str, Any]) -> list[Any]:
    """Map material dict to a row matching TEMPLATE_HEADERS order (operacao=E, codigo_material=id_erp)."""
    return [
        "E",  # operacao
        m.get("id_erp") or "",
        m.get("description") or "",
        m.get("pdm_code") or "",
        m.get("status") or "Ativo",
        m.get("unit_of_measure") or "",
        m.get("material_type") or "",
        "",  # industry_sector
        "",  # base_unit
        m.get("gross_weight"),
        m.get("net_weight"),
        "",  # weight_unit
        "",  # volume
        "",  # volume_unit
        m.get("lead_time"),
        m.get("min_stock"),
        m.get("max_stock"),
        m.get("standard_price"),
        "",  # price_unit
        "",  # currency
        m.get("ncm") or "",
        m.get("cfop") or "",
        m.get("origin") or "",
    ]


def build_export_xlsx(materials: list[dict[str, Any]]) -> BytesIO:
    """Build Excel export with header and one row per material. Same format as import template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Materiais"

    # Row 1: header (same style as template)
    for col_idx, header in enumerate(TEMPLATE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Rows 2+: one per material
    for row_idx, m in enumerate(materials, start=2):
        row_vals = _material_to_export_row(m)
        for col_idx, val in enumerate(row_vals, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Adjust column widths
    for col_idx in range(1, len(TEMPLATE_HEADERS) + 1):
        max_len = len(str(TEMPLATE_HEADERS[col_idx - 1])) + 2
        for r in range(2, len(materials) + 2):
            v = ws.cell(row=r, column=col_idx).value
            if v is not None:
                max_len = max(max_len, len(str(v)) + 2)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_len, 50))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_template_xlsx() -> BytesIO:
    """Build the Excel template and return as BytesIO."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Materiais"

    # Linha 1: cabeçalho
    for col_idx, header in enumerate(TEMPLATE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Linha 2: exemplo
    for col_idx, value in enumerate(EXAMPLE_ROW, start=1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.fill = EXAMPLE_FILL

    # Linhas 3-10: vazias
    for row in range(3, 11):
        for col_idx in range(1, len(TEMPLATE_HEADERS) + 1):
            ws.cell(row=row, column=col_idx, value="")

    # Ajustar largura das colunas
    for col_idx in range(1, len(TEMPLATE_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            12, len(str(TEMPLATE_HEADERS[col_idx - 1])) + 2
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _cell_value(cell) -> str | float | int | None:
    """Get cell value, empty string as None."""
    val = cell.value
    if val is None:
        return None
    if isinstance(val, str) and val.strip() == "":
        return None
    return val


def _safe_float(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v).strip().replace(",", ".")
        return float(s) if s else None
    except (ValueError, TypeError):
        return None


def _safe_int(v: Any) -> int | None:
    if v is None:
        return None
    if isinstance(v, int):
        return v
    f = _safe_float(v)
    return int(f) if f is not None else None


def parse_and_validate_excel(
    file_content: bytes,
    db,
    MaterialDatabaseORM,
    PDMOrm,
) -> dict:
    """
    Parse Excel file and validate. Returns the response dict.
    Does NOT persist — caller handles dry_run and commit.
    """
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(file_content), read_only=False, data_only=True)
    if "Materiais" not in wb.sheetnames:
        return {"_error": "Aba 'Materiais' não encontrada no arquivo."}

    ws = wb["Materiais"]
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if not header_row:
        return {"_error": "Cabeçalho não encontrado."}

    headers = [str(h).strip().lower() if h else "" for h in header_row[0]]
    col_map: dict[str, int] = {}
    for idx, h in enumerate(headers):
        if h:
            col_map[h] = idx + 1

    # Validar presença de colunas essenciais
    required_cols = ["operacao", "descricao", "pdm_code"]
    for rc in required_cols:
        if rc not in col_map:
            return {"_error": f"Coluna obrigatória '{rc}' não encontrada no cabeçalho."}

    # PDMs existentes (internal_code)
    pdm_codes = {r.internal_code for r in db.query(PDMOrm).all()}

    rows_result: list[dict] = []
    total_rows = 0
    valid_count = 0
    error_count = 0
    warning_count = 0

    for row_num in range(2, ws.max_row + 1):
        row_data: dict[str, Any] = {}
        for col_name, col_idx in col_map.items():
            cell = ws.cell(row=row_num, column=col_idx)
            row_data[col_name] = _cell_value(cell)

        # Pular linhas completamente vazias
        if all(v is None for v in row_data.values()):
            continue

        total_rows += 1
        errors: list[str] = []
        warnings: list[str] = []
        operacao_raw = row_data.get("operacao")
        operacao = str(operacao_raw).strip().upper() if operacao_raw else ""

        # a. operacao deve ser C ou E
        if operacao not in ("C", "E"):
            errors.append("operacao deve ser 'C' (Criar) ou 'E' (Editar)")
            rows_result.append({
                "row_number": row_num,
                "operacao": operacao or str(operacao_raw),
                "codigo_material": row_data.get("codigo_material"),
                "descricao": row_data.get("descricao"),
                "status": "error",
                "errors": errors,
                "warnings": warnings,
                "data": row_data,
            })
            error_count += 1
            continue

        # b. Se E: codigo_material obrigatório
        codigo = row_data.get("codigo_material")
        codigo_str = str(codigo).strip() if codigo is not None else ""
        if operacao == "E" and not codigo_str:
            errors.append("codigo_material é obrigatório para operacao E (Editar)")

        # c. Se E: material deve existir
        material_row = None
        if operacao == "E" and codigo_str:
            material_row = db.query(MaterialDatabaseORM).filter(
                MaterialDatabaseORM.id_erp == codigo_str
            ).first()
            if not material_row:
                errors.append(f"Material com código '{codigo_str}' não encontrado no banco")

        # d. Se C: descricao obrigatória
        descricao = row_data.get("descricao")
        descricao_str = str(descricao).strip() if descricao is not None else ""
        if operacao == "C" and not descricao_str:
            errors.append("descricao é obrigatória para operacao C (Criar)")

        # e. Se C: pdm_code obrigatório e deve existir
        pdm_code_val = row_data.get("pdm_code")
        pdm_code_str = str(pdm_code_val).strip() if pdm_code_val is not None else ""
        if operacao == "C":
            if not pdm_code_str:
                errors.append("pdm_code é obrigatório para operacao C (Criar)")
            elif pdm_code_str not in pdm_codes:
                errors.append(f"PDM '{pdm_code_str}' não encontrado em pdm_templates")

        # f. status se preenchido deve ser válido
        status_val = row_data.get("status")
        status_str = str(status_val).strip() if status_val else ""
        if status_str and status_str not in ("Ativo", "Bloqueado", "Obsoleto"):
            warnings.append(f"status '{status_str}' inválido; use Ativo, Bloqueado ou Obsoleto")

        # g. Campos numéricos
        for num_col in ("gross_weight", "net_weight", "lead_time", "min_stock", "max_stock", "standard_price"):
            v = row_data.get(num_col)
            if v is not None and str(v).strip():
                if num_col == "lead_time":
                    parsed = _safe_int(v)
                else:
                    parsed = _safe_float(v)
                if parsed is None:
                    warnings.append(f"{num_col} deve ser numérico")

        status_out = "error" if errors else ("warning" if warnings else "ok")
        if errors:
            error_count += 1
        elif warnings:
            warning_count += 1
        else:
            valid_count += 1

        rows_result.append({
            "row_number": row_num,
            "operacao": operacao,
            "codigo_material": codigo_str if codigo_str else None,
            "descricao": descricao_str or None,
            "status": status_out,
            "errors": errors,
            "warnings": warnings,
            "data": row_data,
        })

    return {
        "total_rows": total_rows,
        "valid_rows": valid_count,
        "error_rows": error_count,
        "warning_rows": warning_count,
        "rows": rows_result,
    }


def _row_to_create_kwargs(row_data: dict, id_erp: str, pdm_name: str) -> dict:
    """Build kwargs for MaterialDatabaseORM create from validated row data."""
    status_val = row_data.get("status")
    status_str = str(status_val).strip() if status_val else "Ativo"
    if status_str not in ("Ativo", "Bloqueado", "Obsoleto"):
        status_str = "Ativo"
    kwargs: dict[str, Any] = {
        "id_erp": id_erp,
        "description": str(row_data.get("descricao", "")).strip() or "Sem descrição",
        "status": status_str,
        "pdm_code": str(row_data.get("pdm_code", "")).strip() or None,
        "pdm_name": pdm_name,
        "source": "bulk_import",
    }
    for col, model_key in COL_TO_MODEL.items():
        if col in ("descricao", "pdm_code", "status"):
            continue
        v = row_data.get(col)
        if v is None or (isinstance(v, str) and v.strip() == ""):
            continue
        if col == "lead_time":
            parsed = _safe_int(v)
        elif col in ("gross_weight", "net_weight", "min_stock", "max_stock", "standard_price"):
            parsed = _safe_float(v)
        else:
            parsed = str(v).strip() if v else None
        if parsed is not None:
            kwargs[model_key] = parsed
    return kwargs


def _row_to_update_kwargs(row_data: dict) -> dict:
    """Build kwargs for MaterialDatabaseORM update (only non-empty fields)."""
    kwargs: dict[str, Any] = {}
    for col, model_key in COL_TO_MODEL.items():
        v = row_data.get(col)
        if v is None or (isinstance(v, str) and v.strip() == ""):
            continue
        if col == "lead_time":
            parsed = _safe_int(v)
        elif col in ("gross_weight", "net_weight", "min_stock", "max_stock", "standard_price"):
            parsed = _safe_float(v)
        else:
            parsed = str(v).strip() if v else None
        if parsed is not None:
            kwargs[model_key] = parsed
    return kwargs
