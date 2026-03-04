"""
Bulk import of PDM and attributes via Excel.

Used by:
  GET  /api/pdm/import-template
  POST /api/pdm/import
  GET  /api/pdm/export
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from bulk_import import HEADER_FILL, HEADER_FONT, EXAMPLE_FILL

PDM_HEADERS = ["operacao", "pdm_code", "nome", "descricao", "ativo"]

ATTR_HEADERS = [
    "operacao",
    "pdm_code",
    "atributo_key",
    "label",
    "tipo",
    "obrigatorio",
    "ordem",
    "opcoes",
]

PDM_EXAMPLE_ROW = [
    "C",
    "PDM-EX-001",
    "Exemplo PDM",
    "Descrição do PDM",
    "Sim",
]

ATTR_EXAMPLE_ROW_1 = [
    "C",
    "PDM-EX-001",
    "diametro",
    "Diâmetro",
    "text",
    "Sim",
    1,
    "",
]

ATTR_EXAMPLE_ROW_2 = [
    "C",
    "PDM-EX-001",
    "material_base",
    "Material Base",
    "select",
    "Não",
    2,
    "Aço;Alumínio;Inox",
]

VALID_PDM_OPS = ("C", "E")
VALID_ATTR_OPS = ("C", "E", "D")
VALID_TIPOS = ("text", "number", "select", "date", "textarea")
VALID_ATIVO = ("Sim", "Não")
VALID_OBRIGATORIO = ("Sim", "Não")


def build_pdm_template_xlsx() -> BytesIO:
    """Build Excel template with PDM and Atributos sheets."""
    wb = Workbook()
    ws_pdm = wb.active
    ws_pdm.title = "PDM"

    for col_idx, header in enumerate(PDM_HEADERS, start=1):
        cell = ws_pdm.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, val in enumerate(PDM_EXAMPLE_ROW, start=1):
        cell = ws_pdm.cell(row=2, column=col_idx, value=val)
        cell.fill = EXAMPLE_FILL
    for row in range(3, 8):
        for col_idx in range(1, len(PDM_HEADERS) + 1):
            ws_pdm.cell(row=row, column=col_idx, value="")
    for col_idx in range(1, len(PDM_HEADERS) + 1):
        ws_pdm.column_dimensions[get_column_letter(col_idx)].width = max(
            12, len(str(PDM_HEADERS[col_idx - 1])) + 2
        )

    ws_attr = wb.create_sheet("Atributos", 1)
    for col_idx, header in enumerate(ATTR_HEADERS, start=1):
        cell = ws_attr.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, val in enumerate(ATTR_EXAMPLE_ROW_1, start=1):
        cell = ws_attr.cell(row=2, column=col_idx, value=val)
        cell.fill = EXAMPLE_FILL
    for col_idx, val in enumerate(ATTR_EXAMPLE_ROW_2, start=1):
        cell = ws_attr.cell(row=3, column=col_idx, value=val)
        cell.fill = EXAMPLE_FILL
    for row in range(4, 12):
        for col_idx in range(1, len(ATTR_HEADERS) + 1):
            ws_attr.cell(row=row, column=col_idx, value="")
    for col_idx in range(1, len(ATTR_HEADERS) + 1):
        ws_attr.column_dimensions[get_column_letter(col_idx)].width = max(
            12, len(str(ATTR_HEADERS[col_idx - 1])) + 2
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_pdm_export_xlsx(
    pdms: list[dict[str, Any]], attributes: list[dict[str, Any]]
) -> BytesIO:
    """Build Excel export with real PDM and attribute data."""
    wb = Workbook()
    ws_pdm = wb.active
    ws_pdm.title = "PDM"
    for col_idx, header in enumerate(PDM_HEADERS, start=1):
        cell = ws_pdm.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx, p in enumerate(pdms, start=2):
        ws_pdm.cell(row=row_idx, column=1, value="E")
        ws_pdm.cell(row=row_idx, column=2, value=p.get("internal_code") or "")
        ws_pdm.cell(row=row_idx, column=3, value=p.get("name") or "")
        ws_pdm.cell(row=row_idx, column=4, value="")
        ws_pdm.cell(row=row_idx, column=5, value="Sim" if p.get("is_active") else "Não")
    for col_idx in range(1, len(PDM_HEADERS) + 1):
        ws_pdm.column_dimensions[get_column_letter(col_idx)].width = max(12, 14)

    ws_attr = wb.create_sheet("Atributos", 1)
    for col_idx, header in enumerate(ATTR_HEADERS, start=1):
        cell = ws_attr.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx, a in enumerate(attributes, start=2):
        opts = a.get("allowedValues") or []
        opcoes = ";".join(
            ov.get("value", ov) if isinstance(ov, dict) else str(ov) for ov in opts
        )
        tipo = (a.get("dataType") or "text").lower()
        if tipo == "lov":
            tipo = "select"
        elif tipo == "numeric":
            tipo = "number"
        ws_attr.cell(row=row_idx, column=1, value="E")
        ws_attr.cell(row=row_idx, column=2, value=a.get("pdm_code") or "")
        ws_attr.cell(row=row_idx, column=3, value=a.get("name") or a.get("id") or "")
        ws_attr.cell(row=row_idx, column=4, value=a.get("name") or "")
        ws_attr.cell(row=row_idx, column=5, value=tipo)
        ws_attr.cell(row=row_idx, column=6, value="Sim" if a.get("isRequired") else "Não")
        ws_attr.cell(row=row_idx, column=7, value=a.get("order", 0))
        ws_attr.cell(row=row_idx, column=8, value=opcoes)
    for col_idx in range(1, len(ATTR_HEADERS) + 1):
        ws_attr.column_dimensions[get_column_letter(col_idx)].width = max(12, 16)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _cell_value(cell) -> str | int | float | None:
    val = cell.value
    if val is None:
        return None
    if isinstance(val, str) and val.strip() == "":
        return None
    return val


def _safe_int(v: Any) -> int | None:
    if v is None:
        return None
    if isinstance(v, int):
        return v
    try:
        s = str(v).strip().replace(",", ".")
        return int(float(s)) if s else None
    except (ValueError, TypeError):
        return None


def parse_and_validate_pdm_excel(file_content: bytes, db) -> dict:
    """
    Parse and validate PDM Excel file. Returns result for both sheets.
    Does NOT persist — caller handles dry_run and commit.
    """
    from orm_models import PDMOrm

    wb = load_workbook(BytesIO(file_content), read_only=False, data_only=True)
    pdm_codes_in_db = {r.internal_code for r in db.query(PDMOrm).all()}
    pdm_codes_created_in_sheet: set[str] = set()

    result: dict[str, Any] = {"dry_run": True, "pdm": None, "attributes": None}

    # ─── Aba PDM ─────────────────────────────────────────────────────────
    if "PDM" not in wb.sheetnames:
        result["_error"] = "Aba 'PDM' não encontrada no arquivo."
        return result

    ws_pdm = wb["PDM"]
    header_row = list(ws_pdm.iter_rows(min_row=1, max_row=1, values_only=True))
    if not header_row:
        result["_error"] = "Cabeçalho da aba PDM não encontrado."
        return result

    headers_pdm = [str(h).strip().lower() if h else "" for h in header_row[0]]
    col_map_pdm: dict[str, int] = {}
    for idx, h in enumerate(headers_pdm):
        if h:
            col_map_pdm[h] = idx + 1

    for required in ["operacao", "pdm_code", "nome"]:
        if required not in col_map_pdm:
            result["_error"] = f"Coluna obrigatória '{required}' não encontrada na aba PDM."
            return result

    pdm_rows_result: list[dict] = []
    pdm_total = 0
    pdm_valid = 0
    pdm_error = 0
    pdm_warning = 0

    for row_num in range(2, ws_pdm.max_row + 1):
        row_data: dict[str, Any] = {}
        for col_name, col_idx in col_map_pdm.items():
            cell = ws_pdm.cell(row=row_num, column=col_idx)
            row_data[col_name] = _cell_value(cell)

        if all(v is None for v in row_data.values()):
            continue

        pdm_total += 1
        errors: list[str] = []
        warnings: list[str] = []
        op = str(row_data.get("operacao", "") or "").strip().upper()
        pdm_code = str(row_data.get("pdm_code") or "").strip()
        nome = str(row_data.get("nome") or "").strip()
        ativo_raw = str(row_data.get("ativo") or "").strip()

        if op not in VALID_PDM_OPS:
            errors.append("operacao deve ser 'C' (Criar) ou 'E' (Editar)")
            pdm_rows_result.append({
                "row_number": row_num,
                "operacao": op or "?",
                "pdm_code": pdm_code or None,
                "nome": nome or None,
                "status": "error",
                "errors": errors,
                "warnings": warnings,
                "data": row_data,
            })
            pdm_error += 1
            continue

        if op == "E" and not pdm_code:
            errors.append("pdm_code é obrigatório para operacao E (Editar)")

        if op == "E" and pdm_code and pdm_code not in pdm_codes_in_db:
            errors.append(f"PDM '{pdm_code}' não encontrado. Para criar um novo PDM use operacao=C.")

        if op == "C" and not nome:
            errors.append("nome é obrigatório para operacao C (Criar)")

        if ativo_raw and ativo_raw not in VALID_ATIVO:
            warnings.append("ativo deve ser 'Sim' ou 'Não'")

        if op == "C" and pdm_code:
            pdm_codes_created_in_sheet.add(pdm_code)

        status_out = "error" if errors else ("warning" if warnings else "ok")
        if errors:
            pdm_error += 1
        elif warnings:
            pdm_warning += 1
        else:
            pdm_valid += 1

        pdm_rows_result.append({
            "row_number": row_num,
            "operacao": op,
            "pdm_code": pdm_code or None,
            "nome": nome or None,
            "status": status_out,
            "errors": errors,
            "warnings": warnings,
            "data": row_data,
        })

    result["pdm"] = {
        "total_rows": pdm_total,
        "valid_rows": pdm_valid,
        "error_rows": pdm_error,
        "warning_rows": pdm_warning,
        "rows": pdm_rows_result,
    }

    # ─── Aba Atributos ───────────────────────────────────────────────────
    if "Atributos" not in wb.sheetnames:
        result["_error"] = "Aba 'Atributos' não encontrada no arquivo."
        return result

    ws_attr = wb["Atributos"]
    header_row_attr = list(ws_attr.iter_rows(min_row=1, max_row=1, values_only=True))
    if not header_row_attr:
        result["_error"] = "Cabeçalho da aba Atributos não encontrado."
        return result

    headers_attr = [str(h).strip().lower() if h else "" for h in header_row_attr[0]]
    col_map_attr: dict[str, int] = {}
    for idx, h in enumerate(headers_attr):
        if h:
            col_map_attr[h] = idx + 1

    valid_pdm_codes = pdm_codes_in_db | pdm_codes_created_in_sheet

    # Build set of (pdm_code, atributo_key) for attributes that exist in DB (for E validation)
    existing_attr_keys: set[tuple[str, str]] = set()
    for r in db.query(PDMOrm).filter(PDMOrm.internal_code.in_(list(pdm_codes_in_db))).all():
        for a in (r.attributes or []):
            if isinstance(a, dict):
                k = str(a.get("id", a.get("name", ""))).strip()
                if k:
                    existing_attr_keys.add((r.internal_code, k))

    attr_rows_result: list[dict] = []
    attr_total = 0
    attr_valid = 0
    attr_error = 0
    attr_warning = 0

    for row_num in range(2, ws_attr.max_row + 1):
        row_data = {}
        for col_name, col_idx in col_map_attr.items():
            cell = ws_attr.cell(row=row_num, column=col_idx)
            row_data[col_name] = _cell_value(cell)

        if all(v is None for v in row_data.values()):
            continue

        attr_total += 1
        errors = []
        warnings = []
        op = str(row_data.get("operacao", "") or "").strip().upper()
        pdm_code = str(row_data.get("pdm_code") or "").strip()
        atributo_key = str(row_data.get("atributo_key") or "").strip()
        tipo_raw = str(row_data.get("tipo") or "").strip().lower()
        obrigatorio_raw = str(row_data.get("obrigatorio") or "").strip()
        ordem_raw = row_data.get("ordem")

        if op not in VALID_ATTR_OPS:
            errors.append("operacao deve ser 'C', 'E' ou 'D'")
            attr_rows_result.append({
                "row_number": row_num,
                "operacao": op or "?",
                "pdm_code": pdm_code or None,
                "atributo_key": atributo_key or None,
                "status": "error",
                "errors": errors,
                "warnings": warnings,
                "data": row_data,
            })
            attr_error += 1
            continue

        if not pdm_code:
            errors.append("pdm_code é obrigatório")

        if pdm_code and pdm_code not in valid_pdm_codes:
            errors.append(f"PDM '{pdm_code}' não encontrado no banco nem sendo criado na aba PDM")

        if op in ("C", "E") and not atributo_key:
            errors.append("atributo_key é obrigatório para operacao C e E")

        if op == "D" and not atributo_key:
            errors.append("operacao=D requer pdm_code e atributo_key para identificar o atributo a remover")

        if op == "E" and pdm_code and atributo_key and (pdm_code, atributo_key) not in existing_attr_keys:
            if pdm_code in valid_pdm_codes:
                errors.append(
                    f"Atributo '{atributo_key}' não encontrado no PDM '{pdm_code}'. "
                    "Para criar use operacao=C ou verifique a key."
                )
            # else pdm_code error already added above

        if tipo_raw and tipo_raw not in VALID_TIPOS:
            warnings.append(f"tipo deve ser um de: {', '.join(VALID_TIPOS)}")

        if obrigatorio_raw and obrigatorio_raw not in VALID_OBRIGATORIO:
            warnings.append("obrigatorio deve ser 'Sim' ou 'Não'")

        if ordem_raw is not None and ordem_raw != "":
            if _safe_int(ordem_raw) is None:
                warnings.append("ordem deve ser numérico")

        status_out = "error" if errors else ("warning" if warnings else "ok")
        if errors:
            attr_error += 1
        elif warnings:
            attr_warning += 1
        else:
            attr_valid += 1

        attr_rows_result.append({
            "row_number": row_num,
            "operacao": op,
            "pdm_code": pdm_code or None,
            "atributo_key": atributo_key or None,
            "status": status_out,
            "errors": errors,
            "warnings": warnings,
            "data": row_data,
        })

    result["attributes"] = {
        "total_rows": attr_total,
        "valid_rows": attr_valid,
        "error_rows": attr_error,
        "warning_rows": attr_warning,
        "rows": attr_rows_result,
    }

    return result
