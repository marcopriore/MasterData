"""
Admin router — Gestão de Usuários e Acessos (v1.8)

Prefix  : /admin
Tags    : ["Admin – Roles", "Admin – Users", "Auth"]

Endpoints
─────────────────────────────────────────────────────────────────────────────
Roles
  GET    /admin/roles                  List all roles with permissions
  GET    /admin/roles/{role_id}        Get a single role
  POST   /admin/roles                  Create a new role
  PATCH  /admin/roles/{role_id}        Update role name / permissions
  DELETE /admin/roles/{role_id}        Delete role (blocked if users assigned)

Users
  GET    /admin/users                  List users (filterable by role / status)
  GET    /admin/users/{user_id}        Get a single user
  POST   /admin/users                  Create user with hashed password
  PUT    /admin/users/{user_id}        Full update of user data and role
  PATCH  /admin/users/{user_id}        Partial update
  PATCH  /admin/users/{user_id}/password   Change password (requires current)
  PATCH  /admin/users/{user_id}/preferences  Update theme / language
  DELETE /admin/users/{user_id}        Soft-delete (deactivate)

Auth
  POST   /admin/auth/login             Credential check → user + role payload
"""

import secrets
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, Body, Depends, File, HTTPException, Query, Request, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session, joinedload

from deps import get_db, get_db_always_raw, get_admin_user, get_current_user, get_current_user_optional, get_user_with_manage_users, get_user_with_view_logs, refresh_with_rls, require_master, set_tenant_in_session
from limiter import limiter
from orm_models import (
    FieldDictionaryORM,
    MaterialRequestORM,
    MaterialDatabaseORM,
    RoleORM,
    TenantORM,
    UserORM,
    WorkflowConfigORM,
    WorkflowHeaderORM,
    SystemLogORM,
)
from constants import ONBOARDING_FIELD_DICT, ONBOARDING_ROLE_DEFS, ONBOARDING_WORKFLOW_STEPS
from email_service import send_welcome_email
from security import create_access_token, hash_password, verify_password
from audit import log_system_event
from bulk_import import HEADER_FILL, HEADER_FONT
from bulk_import_users import (
    build_user_template_xlsx,
    build_user_export_xlsx,
    parse_and_validate_user_excel,
)
from models import (
    LoginRequest,
    RoleCreate,
    RoleUpdate,
    SwitchTenantBody,
    TenantCreate,
    TenantOnboardingRequest,
    TenantUpdate,
    UserCreate,
    UserPasswordChange,
    UserPreferences,
    UserUpdate,
)

router = APIRouter(prefix="/admin", tags=["Admin"])


# ─── Serialisers ──────────────────────────────────────────────────────────────

def _role_to_dict(r: RoleORM) -> dict:
    return {
        "id": r.id,
        "name": r.name,
        "role_type": getattr(r, "role_type", "sistema"),
        "permissions": r.permissions or {},
        "user_count": len(r.users) if r.users is not None else 0,
    }


def _user_to_dict(u: UserORM) -> dict:
    role_name = u.role.name if u.role else None
    is_master = (role_name or "").upper() == "MASTER"
    out: dict = {
        "id": u.id,
        "name": u.name,
        "email": u.email,
        "role_id": u.role_id,
        "role_name": role_name,
        "role_type": getattr(u.role, "role_type", "sistema") if u.role else "sistema",
        "role_permissions": u.role.permissions if u.role else {},
        "is_active": u.is_active,
        "preferences": u.preferences or {"theme": "light", "language": "pt"},
        "created_at": u.created_at.isoformat() if u.created_at else None,
    }
    out["tenant_id"] = getattr(u, "tenant_id", None)
    out["tenant_name"] = u.tenant.name if u.tenant else None
    out["max_description_length"] = getattr(u.tenant, "max_description_length", None) or 40
    out["is_master"] = is_master
    return out


def _load_user(user_id: int, db: Session) -> UserORM:
    """Fetch a user with its role eagerly loaded, or raise 404."""
    row = (
        db.query(UserORM)
        .options(joinedload(UserORM.role))
        .filter(UserORM.id == user_id)
        .first()
    )
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Usuário não encontrado")
    return row


def _load_role(role_id: int, db: Session) -> RoleORM:
    row = (
        db.query(RoleORM)
        .options(joinedload(RoleORM.users))
        .filter(RoleORM.id == role_id)
        .first()
    )
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Perfil não encontrado")
    return row


# ═══════════════════════════════════════════════════════════════════════════════
# ROLES
# ═══════════════════════════════════════════════════════════════════════════════

@router.get(
    "/roles",
    summary="Lista todos os perfis de acesso",
    response_description="Array de perfis com suas permissões e contagem de usuários",
)
def list_roles(
    db: Session = Depends(get_db),
    _: UserORM = Depends(get_current_user),
):
    """
    Retorna todos os perfis cadastrados (ADMIN, SOLICITANTE, TRIAGEM, FISCAL, MASTER, MRP, …)
    com o conjunto de flags de permissão e o número de usuários vinculados.
    """
    rows = (
        db.query(RoleORM)
        .options(joinedload(RoleORM.users))
        .order_by(RoleORM.id)
        .all()
    )
    return [_role_to_dict(r) for r in rows]


@router.get(
    "/roles/{role_id}",
    summary="Detalha um perfil de acesso",
)
def get_role(
    role_id: int,
    db: Session = Depends(get_db),
    _: UserORM = Depends(get_current_user),
):
    return _role_to_dict(_load_role(role_id, db))


@router.post(
    "/roles",
    status_code=status.HTTP_201_CREATED,
    summary="Cria um novo perfil de acesso",
)
def create_role(
    payload: RoleCreate,
    db: Session = Depends(get_db),
    current_user: Optional[UserORM] = Depends(get_current_user_optional),
):
    """
    O `name` é automaticamente convertido para maiúsculas.
    As `permissions` são flags booleanas; omita para usar os padrões (todos False
    exceto `can_submit_request`).
    """
    name_upper = payload.name.strip().upper()
    tenant_id = current_user.tenant_id if current_user else None
    if tenant_id is None:
        raise HTTPException(status_code=401, detail="Autenticação necessária para criar perfil")
    if db.query(RoleORM).filter(RoleORM.tenant_id == tenant_id, RoleORM.name == name_upper).first():
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Perfil '{name_upper}' já existe",
        )
    row = RoleORM(
        tenant_id=tenant_id,
        name=name_upper,
        role_type=payload.role_type,
        permissions=payload.permissions.model_dump(),
    )
    db.add(row)
    db.commit()
    refresh_with_rls(db, row, tenant_id, getattr(current_user, "is_master", False))
    creator = current_user.name if current_user else "Sistema"
    log_system_event(
        db, current_user.id if current_user else None, "roles", "role_created",
        f"Perfil '{name_upper}' criado por {creator}",
        tenant_id,
        is_master=getattr(current_user, "is_master", False) if current_user else False,
    )
    return _role_to_dict(_load_role(row.id, db))


@router.patch(
    "/roles/{role_id}",
    summary="Atualiza nome e/ou permissões de um perfil",
)
def update_role(
    role_id: int,
    payload: RoleUpdate,
    db: Session = Depends(get_db),
    current_user: Optional[UserORM] = Depends(get_current_user_optional),
):
    row = _load_role(role_id, db)
    if payload.role_type is not None:
        row.role_type = payload.role_type
    if payload.name is not None:
        new_name = payload.name.strip().upper()
        conflict = (
            db.query(RoleORM)
            .filter(RoleORM.name == new_name, RoleORM.id != role_id)
            .first()
        )
        if conflict:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Perfil '{new_name}' já existe",
            )
        row.name = new_name
    if payload.permissions is not None:
        row.permissions = payload.permissions.model_dump()
    tid = row.tenant_id
    db.commit()
    refresh_with_rls(db, row, tid, getattr(current_user, "is_master", False) if current_user else False)
    updater = current_user.name if current_user else "Sistema"
    log_system_event(
        db, current_user.id if current_user else None, "roles", "role_updated",
        f"Perfil #{role_id} ({row.name}) atualizado por {updater}",
        row.tenant_id,
        is_master=getattr(current_user, "is_master", False) if current_user else False,
    )
    return _role_to_dict(_load_role(row.id, db))


@router.delete(
    "/roles/{role_id}",
    summary="Remove um perfil de acesso",
)
def delete_role(
    role_id: int,
    db: Session = Depends(get_db),
    _: UserORM = Depends(get_current_user),
):
    """
    Bloqueado se houver usuários vinculados ao perfil.
    Reatribua-os antes de excluir.
    """
    row = _load_role(role_id, db)
    if db.query(UserORM).filter(UserORM.role_id == role_id).count() > 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Não é possível excluir: existem usuários vinculados a este perfil. Reatribua-os primeiro.",
        )
    db.delete(row)
    db.commit()
    return {"ok": True}


# ═══════════════════════════════════════════════════════════════════════════════
# USERS
# ═══════════════════════════════════════════════════════════════════════════════

@router.get(
    "/users",
    summary="Lista todos os usuários com seus perfis",
    response_description="Array de usuários com dados completos de perfil",
)
def list_users(
    role_id: Optional[int] = Query(None, description="Filtrar por perfil"),
    is_active: Optional[bool] = Query(None, description="Filtrar por status (ativo/inativo)"),
    search: Optional[str] = Query(None, description="Busca por nome ou e-mail (case-insensitive)"),
    db: Session = Depends(get_db),
    _: UserORM = Depends(get_current_user),
):
    """
    Retorna todos os usuários com seus respectivos perfis de acesso.
    Suporta filtros opcionais por `role_id`, `is_active` e texto livre (`search`).
    """
    q = db.query(UserORM).options(joinedload(UserORM.role)).order_by(UserORM.id)
    if role_id is not None:
        q = q.filter(UserORM.role_id == role_id)
    if is_active is not None:
        q = q.filter(UserORM.is_active == is_active)
    if search:
        term = f"%{search.strip().lower()}%"
        q = q.filter(
            UserORM.name.ilike(term) | UserORM.email.ilike(term)
        )
    return [_user_to_dict(u) for u in q.all()]


# Import/export must be defined before /users/{user_id} to match correctly
DEFAULT_PASSWORD = "Mudar@1234"  # nosec B105 — senha temporária, alterada no primeiro acesso


@router.get(
    "/users/import-template",
    summary="Download template Excel para importação de usuários",
)
def get_user_import_template(
    _: UserORM = Depends(get_user_with_manage_users),
):
    """Retorna planilha modelo com aba Usuários."""
    buf = build_user_template_xlsx()
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_importacao_usuarios.xlsx"},
    )


@router.post(
    "/users/import",
    summary="Importar usuários em massa via Excel",
)
@limiter.limit("10/minute")
async def import_users(
    request: Request,
    file: UploadFile = File(..., description="Planilha Excel para importação"),
    dry_run: bool = Query(True, description="Se True, apenas valida sem gravar"),
    db: Session = Depends(get_db),
    current_user: UserORM = Depends(get_user_with_manage_users),
):
    """Importa usuários conforme planilha. C (criar) ou E (editar). Nunca exporta ou altera senhas."""
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Arquivo deve ser planilha Excel (.xlsx ou .xls)")
    try:
        content = await file.read()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao ler arquivo: {e}")

    try:
        result = parse_and_validate_user_excel(content, db)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Arquivo inválido ou corrompido: {e}")

    if "_error" in result:
        raise HTTPException(status_code=400, detail=result["_error"])

    if dry_run:
        return result

    user_data = result.get("users", {})
    has_errors = any(r.get("status") == "error" for r in user_data.get("rows", []))
    if has_errors:
        raise HTTPException(
            status_code=422,
            detail={
                "message": "Existem erros críticos nas linhas. Corrija e tente novamente.",
                "users": user_data,
            },
        )

    role_by_name: dict[str, int] = {
        (r.name or "").upper(): r.id for r in db.query(RoleORM).all()
    }
    users_by_email: dict[str, UserORM] = {
        u.email.lower(): u for u in db.query(UserORM).options(joinedload(UserORM.role)).all()
    }

    created = 0
    updated = 0

    for row_info in user_data.get("rows", []):
        if row_info.get("status") == "error":
            continue
        op = row_info.get("operacao")
        data = row_info.get("data", {})
        email = str(data.get("email") or "").strip().lower()
        nome = str(data.get("nome") or "").strip() or email
        perfil_raw = str(data.get("perfil") or "").strip().upper()
        ativo_raw = str(data.get("ativo") or "Sim").strip()
        is_active = ativo_raw.lower() in ("sim", "s", "1", "true")

        role_id = role_by_name.get(perfil_raw)
        if not role_id:
            continue

        if op == "C":
            row = UserORM(
                tenant_id=current_user.tenant_id,
                name=nome,
                email=email,
                hashed_password=hash_password(DEFAULT_PASSWORD),
                role_id=role_id,
                is_active=is_active,
            )
            db.add(row)
            created += 1
        elif op == "E":
            row = users_by_email.get(email)
            if row:
                if nome:
                    row.name = nome
                row.role_id = role_id
                row.is_active = is_active
                updated += 1

    db.commit()

    log_system_event(
        db,
        current_user.id,
        "users",
        "bulk_import",
        f"Importação em massa: {created} criados, {updated} atualizados por {current_user.email}",
        current_user.tenant_id,
        is_master=getattr(current_user, "is_master", False),
    )

    return {"dry_run": False, "created": created, "updated": updated}


@router.get(
    "/users/export",
    summary="Exportar todos os usuários para Excel",
)
def export_users(
    db: Session = Depends(get_db),
    _: UserORM = Depends(get_user_with_manage_users),
):
    """Retorna planilha com todos os usuários. NUNCA exporta senhas."""
    rows = (
        db.query(UserORM)
        .options(joinedload(UserORM.role))
        .order_by(UserORM.id)
        .all()
    )
    users = [
        {
            "email": u.email,
            "name": u.name,
            "role_name": u.role.name if u.role else None,
            "is_active": u.is_active,
        }
        for u in rows
    ]
    buf = build_user_export_xlsx(users)
    today = datetime.now(timezone.utc).strftime("%Y%m%d")
    filename = f"usuarios_export_{today}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get(
    "/users/{user_id}",
    summary="Detalha um usuário",
)
def get_user(
    user_id: int,
    db: Session = Depends(get_db),
    _: UserORM = Depends(get_current_user),
):
    return _user_to_dict(_load_user(user_id, db))


@router.post(
    "/users",
    status_code=status.HTTP_201_CREATED,
    summary="Cria novo usuário com senha criptografada",
)
def create_user(
    payload: UserCreate,
    db: Session = Depends(get_db),
    current_user: Optional[UserORM] = Depends(get_current_user_optional),
):
    """
    A senha é recebida em texto plano e armazenada como hash bcrypt.
    O campo `role_id` deve referenciar um perfil existente.
    """
    tenant_id = current_user.tenant_id if current_user else None
    if tenant_id is None:
        raise HTTPException(status_code=401, detail="Autenticação necessária para criar usuário")
    email_lower = payload.email.lower().strip()
    if db.query(UserORM).filter(UserORM.tenant_id == tenant_id, UserORM.email == email_lower).first():
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"E-mail '{payload.email}' já está cadastrado",
        )
    role = db.query(RoleORM).filter(RoleORM.id == payload.role_id).first()
    if not role:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Perfil não encontrado")

    row = UserORM(
        tenant_id=tenant_id,
        name=payload.name.strip(),
        email=email_lower,
        hashed_password=hash_password(payload.password),
        role_id=payload.role_id,
        is_active=True,
        preferences=payload.preferences.model_dump(),
        created_at=datetime.now(timezone.utc),
    )
    db.add(row)
    db.commit()
    refresh_with_rls(db, row, tenant_id, getattr(current_user, "is_master", False) if current_user else False)
    creator = current_user.name if current_user else "Sistema"
    log_system_event(
        db, current_user.id if current_user else None, "users", "user_created",
        f"Usuário {row.email} criado por {creator}",
        tenant_id,
        is_master=getattr(current_user, "is_master", False) if current_user else False,
    )
    # Garantir contexto RLS para _load_user (log_system_event faz commit)
    set_tenant_in_session(db, tenant_id, is_master=getattr(current_user, "is_master", False) if current_user else False)
    return _user_to_dict(_load_user(row.id, db))


@router.put(
    "/users/{user_id}",
    summary="Atualiza dados e perfil do usuário (substituição completa)",
)
def replace_user(
    user_id: int,
    payload: UserCreate,
    db: Session = Depends(get_db),
    current_user: Optional[UserORM] = Depends(get_current_user_optional),
):
    """
    Substitui todos os campos editáveis do usuário.
    Para atualizações parciais use `PATCH /admin/users/{user_id}`.
    A senha é re-hasheada se fornecida.
    """
    row = _load_user(user_id, db)

    # E-mail uniqueness check (excluding self)
    conflict = (
        db.query(UserORM)
        .filter(UserORM.email == payload.email.lower().strip(), UserORM.id != user_id)
        .first()
    )
    if conflict:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="E-mail já está em uso por outro usuário",
        )

    role = db.query(RoleORM).filter(RoleORM.id == payload.role_id).first()
    if not role:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Perfil não encontrado")

    row.name = payload.name.strip()
    row.email = payload.email.lower().strip()
    row.hashed_password = hash_password(payload.password)
    row.role_id = payload.role_id
    row.preferences = payload.preferences.model_dump()
    tid = row.tenant_id
    db.commit()
    refresh_with_rls(db, row, tid, getattr(current_user, "is_master", False) if current_user else False)
    updater = current_user.name if current_user else "Sistema"
    log_system_event(
        db, current_user.id if current_user else None, "users", "user_updated",
        f"Usuário #{user_id} ({row.email}) atualizado por {updater}",
        row.tenant_id,
        is_master=getattr(current_user, "is_master", False) if current_user else False,
    )
    return _user_to_dict(_load_user(row.id, db))


@router.patch(
    "/users/{user_id}",
    summary="Atualização parcial de dados e/ou perfil",
)
def update_user(
    user_id: int,
    payload: UserUpdate,
    db: Session = Depends(get_db),
    current_user: Optional[UserORM] = Depends(get_current_user_optional),
):
    """
    Todos os campos são opcionais. Apenas os campos enviados são alterados.
    """
    row = _load_user(user_id, db)

    if payload.name is not None:
        row.name = payload.name.strip()

    if payload.email is not None:
        new_email = payload.email.lower().strip()
        conflict = (
            db.query(UserORM)
            .filter(UserORM.email == new_email, UserORM.id != user_id)
            .first()
        )
        if conflict:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="E-mail já está em uso por outro usuário",
            )
        row.email = new_email

    if payload.role_id is not None:
        role = db.query(RoleORM).filter(RoleORM.id == payload.role_id).first()
        if not role:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Perfil não encontrado")
        row.role_id = payload.role_id

    if payload.is_active is not None:
        row.is_active = payload.is_active

    if payload.preferences is not None:
        row.preferences = payload.preferences.model_dump()

    tid = row.tenant_id
    db.commit()
    refresh_with_rls(db, row, tid, getattr(current_user, "is_master", False) if current_user else False)
    updater = current_user.name if current_user else "Sistema"
    log_system_event(
        db, current_user.id if current_user else None, "users", "user_updated",
        f"Usuário #{user_id} ({row.email}) atualizado por {updater}",
        row.tenant_id,
        is_master=getattr(current_user, "is_master", False) if current_user else False,
    )
    return _user_to_dict(_load_user(row.id, db))


@router.patch(
    "/users/{user_id}/password",
    summary="Altera a senha do usuário",
)
def change_password(
    user_id: int,
    payload: UserPasswordChange,
    db: Session = Depends(get_db),
    current_user: UserORM = Depends(get_current_user),
):
    """
    Usuário comum: requer a senha atual para confirmar a identidade.
    MASTER e ADMIN: podem redefinir senha de qualquer usuário sem informar a senha atual.
    """
    row = _load_user(user_id, db)

    is_admin_reset = current_user.role and (current_user.role.name or "").upper() in ("MASTER", "ADMIN")

    if not is_admin_reset:
        if not payload.current_password or not verify_password(payload.current_password, row.hashed_password):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Senha atual incorreta",
            )

    row.hashed_password = hash_password(payload.new_password)
    db.commit()
    return {"ok": True, "message": "Senha alterada com sucesso"}


@router.patch(
    "/users/{user_id}/preferences",
    summary="Atualiza preferências de interface (tema e idioma)",
)
def update_preferences(
    user_id: int,
    payload: UserPreferences,
    db: Session = Depends(get_db),
    _: UserORM = Depends(get_current_user),
):
    row = db.query(UserORM).filter(UserORM.id == user_id).first()
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Usuário não encontrado")
    row.preferences = payload.model_dump()
    db.commit()
    return {"ok": True, "preferences": row.preferences}


@router.delete(
    "/users/{user_id}",
    summary="Desativa um usuário (soft-delete)",
)
def delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    _: UserORM = Depends(get_current_user),
):
    """
    Não remove o registro do banco — apenas define `is_active = false`.
    Isso preserva o histórico de solicitações vinculadas ao usuário.
    Use `PATCH /admin/users/{id}` com `is_active: true` para reativar.
    """
    row = db.query(UserORM).filter(UserORM.id == user_id).first()
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Usuário não encontrado")
    row.is_active = False
    db.commit()
    return {"ok": True, "message": "Usuário desativado"}


# ═══════════════════════════════════════════════════════════════════════════════
# AUTH
# ═══════════════════════════════════════════════════════════════════════════════

@router.post(
    "/auth/login",
    summary="Autenticação por e-mail e senha",
    response_description="Dados do usuário autenticado com perfil e permissões",
)
@limiter.limit("10/minute")
def login(
    request: Request,
    payload: LoginRequest,
    db: Session = Depends(get_db_always_raw),
):
    """
    Valida as credenciais e retorna os dados completos do usuário,
    incluindo o perfil e todas as flags de permissão.

    > **Nota:** JWT será adicionado na v1.9. Por ora o frontend deve
    > armazenar o objeto `user` em contexto/sessão.
    """
    row = (
        db.query(UserORM)
        .options(joinedload(UserORM.role), joinedload(UserORM.tenant))
        .filter(UserORM.email == payload.email.lower().strip())
        .first()
    )
    if not row or not verify_password(payload.password, row.hashed_password):
        _tid = row.tenant_id if row else 1
        log_system_event(
            db, None, "auth", "login_failed",
            f"Tentativa de login falhou para {payload.email}",
            _tid,
        )
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="E-mail ou senha inválidos",
        )
    if not row.is_active:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Conta desativada. Entre em contato com o administrador.",
        )
    if not row.tenant or not row.tenant.is_active:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Sua empresa não está ativa no sistema. Entre em contato com o suporte.",
        )
    role_name = row.role.name if row.role else "SOLICITANTE"
    role_type = getattr(row.role, "role_type", "sistema") if row.role else "sistema"
    permissions = row.role.permissions if row.role else {}
    is_master = (role_name or "").upper() == "MASTER"
    log_system_event(
        db, row.id, "auth", "login",
        f"Login realizado com sucesso: {row.email}",
        row.tenant_id,
        is_master=is_master,
    )
    return {
        "ok": True,
        "user": _user_to_dict(row),
        "access_token": create_access_token(
            row.id,
            role_name,
            role_type,
            permissions,
            tenant_id=row.tenant_id,
            is_master=is_master,
        ),
    }


@router.get(
    "/auth/me",
    summary="Retorna dados do usuário autenticado",
)
def auth_me(current_user: UserORM = Depends(get_current_user)):
    """
    Retorna o usuário atual com tenant_id, tenant_name e is_master.
    Requer token JWT.
    """
    return _user_to_dict(current_user)


@router.post(
    "/auth/switch-tenant",
    summary="Chaveia contexto de tenant (apenas MASTER)",
)
@limiter.limit("30/minute")
def switch_tenant(
    request: Request,
    payload: SwitchTenantBody,
    db: Session = Depends(get_db_always_raw),
    current_user: UserORM = Depends(require_master),
):
    """
    Gera novo JWT com tenant_id solicitado, mantendo is_master=True.
    O MASTER continua MASTER mas entra no contexto do tenant.
    """
    tenant = db.query(TenantORM).filter(
        TenantORM.id == payload.tenant_id,
        TenantORM.is_active,
    ).first()
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant não encontrado ou inativo")
    role_name = current_user.role.name if current_user.role else "MASTER"
    role_type = getattr(current_user.role, "role_type", "sistema") if current_user.role else "sistema"
    permissions = current_user.role.permissions if current_user.role else {}
    token = create_access_token(
        current_user.id,
        role_name,
        role_type,
        permissions,
        tenant_id=payload.tenant_id,
        is_master=True,
        master_viewing=True,
        original_tenant_id=current_user.tenant_id,
    )
    return {"access_token": token, "tenant_id": tenant.id, "tenant_name": tenant.name}


@router.get(
    "/auth/switch-tenant/back",
    summary="Retorna ao tenant próprio do MASTER",
)
@limiter.limit("30/minute")
def switch_tenant_back(
    request: Request,
    db: Session = Depends(get_db_always_raw),
    current_user: UserORM = Depends(require_master),
):
    """Gera JWT com tenant_id original do MASTER."""
    role_name = current_user.role.name if current_user.role else "MASTER"
    role_type = getattr(current_user.role, "role_type", "sistema") if current_user.role else "sistema"
    permissions = current_user.role.permissions if current_user.role else {}
    token = create_access_token(
        current_user.id,
        role_name,
        role_type,
        permissions,
        tenant_id=current_user.tenant_id,
        is_master=True,
        master_viewing=False,
    )
    tenant = current_user.tenant
    return {
        "access_token": token,
        "tenant_id": current_user.tenant_id,
        "tenant_name": tenant.name if tenant else None,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# TENANTS (MASTER only)
# ═══════════════════════════════════════════════════════════════════════════════

def _tenant_to_dict(t: TenantORM) -> dict:
    return {
        "id": t.id,
        "name": t.name,
        "slug": t.slug,
        "is_active": t.is_active,
        "max_description_length": getattr(t, "max_description_length", None) or 40,
        "created_at": t.created_at.isoformat() if t.created_at else None,
    }


@router.get("/tenants", summary="Lista todos os tenants")
def list_tenants(
    db: Session = Depends(get_db_always_raw),
    _: UserORM = Depends(require_master),
):
    rows = db.query(TenantORM).order_by(TenantORM.name.asc()).all()
    result = []
    for r in rows:
        d = _tenant_to_dict(r)
        d["users_count"] = db.query(func.count(UserORM.id)).filter(UserORM.tenant_id == r.id).scalar() or 0
        d["materials_count"] = (
            db.query(func.count(MaterialDatabaseORM.id)).filter(MaterialDatabaseORM.tenant_id == r.id).scalar() or 0
        )
        result.append(d)
    return result


@router.post(
    "/tenants/onboarding",
    status_code=status.HTTP_201_CREATED,
    summary="Onboarding completo: tenant + admin + roles + workflow + field dictionary",
)
@limiter.limit("5/minute")
def onboard_tenant(
    request: Request,
    body: TenantOnboardingRequest,
    db: Session = Depends(get_db_always_raw),
    _: UserORM = Depends(require_master),
):
    """
    Cria tenant completo: tenant, roles, admin, workflow padrão, field dictionary.
    Envia email de boas-vindas se SMTP configurado.
    """
    try:
        slug = body.tenant_slug.strip().lower()
        if db.query(TenantORM).filter(TenantORM.slug == slug).first():
            raise HTTPException(status_code=400, detail="Slug já existe")

        email_lower = body.admin_email.strip().lower()
        if db.query(UserORM).filter(UserORM.email == email_lower).first():
            raise HTTPException(status_code=400, detail="E-mail já cadastrado em outro tenant")

        tenant = TenantORM(name=body.tenant_name.strip(), slug=slug, is_active=True)
        db.add(tenant)
        db.flush()

        set_tenant_in_session(db, tenant.id, is_master=False)

        for rdef in ONBOARDING_ROLE_DEFS:
            r = RoleORM(
                tenant_id=tenant.id,
                name=rdef["name"],
                role_type=rdef["role_type"],
                permissions=rdef["permissions"],
            )
            db.add(r)
        db.flush()

        role_admin = db.query(RoleORM).filter(
            RoleORM.tenant_id == tenant.id, RoleORM.name == "ADMIN"
        ).first()

        if body.temp_password and len(body.temp_password) >= 6:
            temp_password = body.temp_password
        else:
            temp_password = secrets.token_urlsafe(10) + "A1!"

        admin_user = UserORM(
            name=body.admin_name.strip(),
            email=email_lower,
            hashed_password=hash_password(temp_password),
            role_id=role_admin.id,
            tenant_id=tenant.id,
            is_active=True,
            preferences={"theme": "light", "language": "pt"},
            created_at=datetime.now(timezone.utc),
        )
        db.add(admin_user)
        db.flush()

        wh = WorkflowHeaderORM(
            tenant_id=tenant.id,
            name="Fluxo Padrão de Cadastro",
            description="Fluxo padrão: Central de Cadastro → Compras → MRP → Fiscal → Contabilidade → Finalizado",
            is_active=True,
        )
        db.add(wh)
        db.flush()

        for step in ONBOARDING_WORKFLOW_STEPS:
            db.add(WorkflowConfigORM(
                tenant_id=tenant.id,
                workflow_id=wh.id,
                step_name=step["step_name"],
                status_key=step["status_key"],
                order=step["order"],
                is_active=True,
            ))

        for entry in ONBOARDING_FIELD_DICT:
            fd = FieldDictionaryORM(
                tenant_id=tenant.id,
                field_name=entry["field_name"],
                field_label=entry["field_label"],
                sap_field=entry.get("sap_field"),
                sap_view=entry["sap_view"],
                field_type=entry["field_type"],
                options=entry.get("options"),
                responsible_role=entry["responsible_role"],
                is_required=entry["is_required"],
                is_active=True,
                display_order=entry["display_order"],
            )
            db.add(fd)

        db.commit()

        email_sent = send_welcome_email(
            to=email_lower,
            admin_name=body.admin_name.strip(),
            tenant_name=tenant.name,
            temp_password=temp_password,
        )

        return {
            "success": True,
            "tenant_id": tenant.id,
            "tenant_name": tenant.name,
            "admin_email": email_lower,
            "email_sent": email_sent,
            "message": "Tenant criado com sucesso.",
        }

    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/tenants", status_code=status.HTTP_201_CREATED, summary="Cria novo tenant")
def create_tenant(
    payload: TenantCreate,
    db: Session = Depends(get_db_always_raw),
    current_user: UserORM = Depends(require_master),
):
    slug = payload.slug.strip().lower()
    if db.query(TenantORM).filter(TenantORM.slug == slug).first():
        raise HTTPException(status_code=400, detail="Slug já existe")
    row = TenantORM(name=payload.name.strip(), slug=slug, is_active=True)
    db.add(row)
    db.commit()
    tid = current_user.tenant_id if current_user else -1
    refresh_with_rls(db, row, tid, getattr(current_user, "is_master", False) if current_user else True)
    return _tenant_to_dict(row)


@router.patch("/tenants/{tenant_id}", summary="Atualiza tenant")
def update_tenant(
    tenant_id: int,
    payload: TenantUpdate,
    db: Session = Depends(get_db_always_raw),
    current_user: UserORM = Depends(require_master),
):
    row = db.query(TenantORM).filter(TenantORM.id == tenant_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Tenant não encontrado")
    if payload.is_active is False and current_user.tenant_id == tenant_id:
        raise HTTPException(
            status_code=400,
            detail="Não é possível desativar o tenant do próprio usuário Master.",
        )
    if payload.name is not None:
        row.name = payload.name.strip()
    if payload.slug is not None:
        new_slug = payload.slug.strip().lower()
        conflict = db.query(TenantORM).filter(
            TenantORM.slug == new_slug,
            TenantORM.id != tenant_id,
        ).first()
        if conflict:
            raise HTTPException(status_code=400, detail="Slug já existe")
        row.slug = new_slug
    if payload.is_active is not None:
        row.is_active = payload.is_active
    tid = current_user.tenant_id
    db.commit()
    refresh_with_rls(db, row, tid, getattr(current_user, "is_master", False))
    return _tenant_to_dict(row)


@router.patch("/tenants/{tenant_id}/settings", summary="Atualiza configurações do tenant")
def update_tenant_settings(
    tenant_id: int,
    payload: dict = Body(...),
    db: Session = Depends(get_db_always_raw),
    current_user: UserORM = Depends(get_admin_user),
):
    """
    Atualiza configurações do tenant.
    Payload: { "max_description_length": 40 }
    Apenas MASTER ou ADMIN.
    """
    tenant = db.query(TenantORM).filter(TenantORM.id == tenant_id).first()
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant não encontrado")

    if "max_description_length" in payload:
        val = int(payload["max_description_length"])
        if val < 10 or val > 200:
            raise HTTPException(
                status_code=400,
                detail="Limite deve ser entre 10 e 200 caracteres",
            )
        tenant.max_description_length = val

    db.commit()
    db.refresh(tenant)
    return {
        "id": tenant.id,
        "name": tenant.name,
        "max_description_length": tenant.max_description_length or 40,
    }


@router.get("/tenants/{tenant_id}/stats", summary="Estatísticas do tenant")
def get_tenant_stats(
    tenant_id: int,
    db: Session = Depends(get_db_always_raw),
    _: UserORM = Depends(require_master),
):
    row = db.query(TenantORM).filter(TenantORM.id == tenant_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Tenant não encontrado")
    users_count = db.query(func.count(UserORM.id)).filter(UserORM.tenant_id == tenant_id).scalar() or 0
    materials_count = db.query(func.count(MaterialDatabaseORM.id)).filter(
        MaterialDatabaseORM.tenant_id == tenant_id
    ).scalar() or 0
    requests_count = db.query(func.count(MaterialRequestORM.id)).filter(
        MaterialRequestORM.tenant_id == tenant_id
    ).scalar() or 0
    pdm_count = (
        db.query(func.count(func.distinct(MaterialRequestORM.pdm_id)))
        .filter(MaterialRequestORM.tenant_id == tenant_id)
        .scalar() or 0
    )
    return {
        "users_count": users_count,
        "materials_count": materials_count,
        "requests_count": requests_count,
        "pdm_count": pdm_count,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# SYSTEM LOGS (ADMIN only)
# ═══════════════════════════════════════════════════════════════════════════════

LOG_EXPORT_HEADERS = ["Data/Hora", "Usuário", "Categoria", "Ação", "Detalhe", "IP"]


def _build_logs_export_xlsx(rows: list) -> bytes:
    """Build Excel with logs. Uses HEADER_FILL, HEADER_FONT from bulk_import."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Logs do Sistema"

    for col_idx, header in enumerate(LOG_EXPORT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, r in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=r.get("created_at") or "")
        ws.cell(row=row_idx, column=2, value=r.get("user_name") or "Sistema")
        ws.cell(row=row_idx, column=3, value=r.get("category") or "")
        ws.cell(row=row_idx, column=4, value=r.get("action") or "")
        detail_cell = ws.cell(row=row_idx, column=5, value=r.get("description") or "")
        detail_cell.alignment = Alignment(wrap_text=True)
        ws.cell(row=row_idx, column=6, value=r.get("ip_address") or "")

    widths = [18, 20, 12, 18, 50, 16]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@router.get(
    "/logs/export",
    summary="Exportar logs do sistema para Excel",
)
def export_logs(
    category: Optional[str] = Query(None, description="Filtrar por categoria"),
    user_id: Optional[int] = Query(None, description="Filtrar por user_id"),
    from_date: Optional[str] = Query(None, alias="from", description="Data inicial ISO"),
    to_date: Optional[str] = Query(None, alias="to", description="Data final ISO"),
    db: Session = Depends(get_db),
    _: UserORM = Depends(get_user_with_view_logs),
):
    """Exporta logs com os mesmos filtros de GET /admin/logs. Sem paginação."""
    q = db.query(SystemLogORM).options(joinedload(SystemLogORM.user))
    if category:
        q = q.filter(SystemLogORM.category == category)
    if user_id is not None:
        q = q.filter(SystemLogORM.user_id == user_id)
    if from_date:
        try:
            from_dt = datetime.fromisoformat(from_date.replace("Z", "+00:00"))
            q = q.filter(SystemLogORM.created_at >= from_dt)
        except ValueError:
            pass
    if to_date:
        try:
            to_dt = datetime.fromisoformat(to_date.replace("Z", "+00:00"))
            q = q.filter(SystemLogORM.created_at <= to_dt)
        except ValueError:
            pass
    q = q.order_by(SystemLogORM.created_at.desc())
    rows = q.all()
    data = [
        {
            "created_at": r.created_at.isoformat() if r.created_at else None,
            "user_name": r.user.name if r.user else None,
            "category": r.category,
            "action": r.action,
            "description": r.description,
            "ip_address": r.ip_address,
        }
        for r in rows
    ]
    from io import BytesIO
    content = _build_logs_export_xlsx(data)
    today = datetime.now(timezone.utc).strftime("%Y%m%d")
    filename = f"logs_export_{today}.xlsx"
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get(
    "/logs",
    summary="Lista logs do sistema (apenas ADMIN)",
)
def list_system_logs(
    category: Optional[str] = Query(None, description="Filtrar por categoria"),
    user_id: Optional[int] = Query(None, description="Filtrar por user_id"),
    from_date: Optional[str] = Query(None, alias="from", description="Data inicial ISO (ex: 2026-01-01)"),
    to_date: Optional[str] = Query(None, alias="to", description="Data final ISO (ex: 2026-03-01)"),
    page: int = Query(1, ge=1, description="Página"),
    limit: int = Query(50, ge=1, le=200, description="Itens por página"),
    db: Session = Depends(get_db),
    _: UserORM = Depends(get_admin_user),
):
    """Retorna lista paginada de logs do sistema."""
    q = db.query(SystemLogORM).options(joinedload(SystemLogORM.user))
    if category:
        q = q.filter(SystemLogORM.category == category)
    if user_id is not None:
        q = q.filter(SystemLogORM.user_id == user_id)
    if from_date:
        try:
            from_dt = datetime.fromisoformat(from_date.replace("Z", "+00:00"))
            q = q.filter(SystemLogORM.created_at >= from_dt)
        except ValueError:
            pass
    if to_date:
        try:
            to_dt = datetime.fromisoformat(to_date.replace("Z", "+00:00"))
            q = q.filter(SystemLogORM.created_at <= to_dt)
        except ValueError:
            pass
    q = q.order_by(SystemLogORM.created_at.desc())
    total = q.count()
    offset = (page - 1) * limit
    rows = q.offset(offset).limit(limit).all()
    return {
        "total": total,
        "page": page,
        "limit": limit,
        "items": [
            {
                "id": r.id,
                "user_name": r.user.name if r.user else None,
                "category": r.category,
                "action": r.action,
                "description": r.description,
                "event_data": r.event_data,
                "ip_address": r.ip_address,
                "created_at": r.created_at.isoformat() if r.created_at else None,
            }
            for r in rows
        ],
    }
